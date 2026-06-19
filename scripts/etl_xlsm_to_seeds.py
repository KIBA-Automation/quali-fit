#!/usr/bin/env python3
"""원천 엑셀(학력_자격증_*.xlsm) → 앱 시드 CSV(Data/*.csv) 변환 ETL.

운영부서가 관리하는 마스터 워크북을 quali-fit이 읽는 시드 CSV로 1회 변환한다.
시드 형식은 db.py 의 SEED_PLAN 과 1:1로 맞춘다(헤더명·FK 순서 동일).

== 단계 ==
이 스크립트는 "1단계"(저위험 마스터)를 채운다:
    01_직원.csv          ← '개인DB' 시트
    02_자격증_마스터.csv  ← '자격증DB' 시트
    04_직원_학력.csv      ← '학력' 시트
나머지 3종은 seed_from_csv()가 6개 CSV를 한 번에 로드하므로 앱이 깨지지 않도록
"헤더만 있는 빈 CSV"로 생성한다(다음 단계에서 채움):
    03_업무코드_마스터.csv
    05_직원_자격증.csv         ← (2단계) '자격증' 시트의 취득/등록/유효기간
    06_업무코드_자격증_매핑.csv ← (3단계) CA~CG 카테고리 매핑

== 사용법 (repo 루트에서) ==
    SRC="/path/to/학력_자격증_26.05.14_v13.xlsm" python scripts/etl_xlsm_to_seeds.py
    # 또는
    python scripts/etl_xlsm_to_seeds.py "/path/to/학력_자격증.xlsm"
    # 그 뒤 DB 적재:
    .venv/Scripts/python -c "import db; db.seed_from_csv()"

== 주의 ==
- 산출되는 Data/*.csv 는 PII/회사 기밀이다. .gitignore 로 제외되며 절대 커밋 금지.
- 이 스크립트(코드)만 버전관리한다.
- 의존성: openpyxl (엑셀 읽기). pandas 불필요(stdlib csv 사용).
"""
import csv
import os
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "Data"

# --- 시트명 (원천 워크북) ---
SHEET_PERSON = "개인DB"
SHEET_CERT = "자격증DB"
SHEET_EDU = "학력"

# --- SEED_PLAN 과 동일한 CSV 헤더 (순서 포함) ---
H_EMPLOYEE = ["직원번호", "이름", "소속", "직책"]
H_CERT = ["코드", "자격증명", "대분류", "중분류", "원가산정검증활용", "자격증내용",
          "수행가능업무", "영향력", "자격유형", "증빙유형", "자격등급구분", "키워드",
          "관련부처", "시행/발급기관"]
H_EDU = ["직원번호", "학력정보번호", "학력", "학위", "학교명", "학부(과)", "전공", "비고"]
H_WORK = ["업무분류코드", "대분류", "중분류", "소분류", "업무구분", "분류기준",
          "관리부서", "책임자", "적용된키워드", "분류근거및설명", "관련지침", "관련법령"]
H_EMP_CERT = ["직원번호", "자격증코드", "취득일", "등록일", "유효기간"]
H_MAP = ["업무분류코드", "자격증코드", "업무관련영향력"]


def s(v) -> str:
    """셀 값을 시드용 문자열로 정규화한다(날짜는 YYYY-MM-DD)."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def importance_to_influence(raw: str) -> str:
    """'중요도' 셀 → influence(1~5 정수 문자열). 비면 중립값 3.

    cert_master.influence 는 INTEGER + CHECK(BETWEEN 1 AND 5) 이라 빈 문자열을
    넣으면 시드가 실패한다. 따라서 항상 유효 정수를 보장한다.
    """
    raw = (raw or "").strip()
    if not raw:
        return "3"
    m = re.search(r"-?\d+", raw)
    if m:
        return str(max(1, min(5, int(m.group()))))
    return {"상": "5", "중": "3", "하": "1",
            "매우높음": "5", "높음": "4", "보통": "3", "낮음": "2"}.get(raw, "3")


def write_csv(name: str, header: list[str], rows: list[list[str]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / name
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  wrote {name:28} rows={len(rows)}")


def load_employees(ws):
    """개인DB → employee 행 + 이름→직원번호 매핑. (헤더 2줄, 데이터 3행~)"""
    rows, name_to_id, dups = [], {}, []
    n = 0
    for r in ws.iter_rows(min_row=3, values_only=True):
        name = s(r[1])          # B: 이름
        if not name:
            continue
        dept, title = s(r[2]), s(r[3])   # C: 소속, D: 직책
        n += 1
        emp_id = f"EMP-{n:03d}"
        rows.append([emp_id, name, dept, title])
        if name in name_to_id:
            dups.append(name)
        else:
            name_to_id[name] = emp_id
    return rows, name_to_id, dups


def load_certs(ws):
    """자격증DB → cert_master 행 + 자격증명→코드 매핑. (헤더 2줄, 데이터 3행~)"""
    rows, name_to_code, dups = [], {}, []
    n = 0
    for r in ws.iter_rows(min_row=3, values_only=True):
        cert_name = s(r[1])     # B: 자격증명
        if not cert_name:
            continue
        ministry = s(r[2])      # C: 관련부처
        issuer = s(r[3])        # D: 시행/발급기관
        license_type = s(r[5])  # F: 구분 (민간/국가/외국...)
        evidence = s(r[6])      # G: 비고 (자격증/등록증)
        importance = s(r[7])    # H: 중요도
        key = s(r[8])           # I: KEY (대표 카테고리 코드, 예: CB-210)
        n += 1
        code = f"CERT-{n:03d}"
        rows.append([
            code, cert_name,
            key, "",                       # 대분류(l1)=KEY, 중분류(l2) 미정
            "", "", "",                    # 원가산정검증활용/자격증내용/수행가능업무
            importance_to_influence(importance),
            license_type, evidence, "", "",  # 자격유형/증빙유형/자격등급구분/키워드
            ministry, issuer,
        ])
        if cert_name in name_to_code:
            dups.append(cert_name)
        else:
            name_to_code[cert_name] = code
    return rows, name_to_code, dups


def load_education(ws, name_to_id):
    """학력 → education 행. 이름으로 직원번호 조인. (헤더 1줄, 데이터 2행~)"""
    rows, missing = [], []
    n = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = s(r[0])          # 이름
        level = s(r[3])         # 학력
        if not name and not level:
            continue
        emp_id = name_to_id.get(name)
        if not emp_id:
            missing.append(name or "(빈 이름)")
            continue
        n += 1
        rows.append([
            emp_id, f"EDU-{n:03d}",
            level,                # 학력 → level
            s(r[4]),              # 학위 → degree
            s(r[5]),              # 학교명 → school
            s(r[6]),              # 학부(과) → faculty
            s(r[7]),              # 전공 → major
            s(r[8]),              # 비고 → note
        ])
    return rows, missing


def main() -> None:
    src = os.environ.get("SRC") or (sys.argv[1] if len(sys.argv) > 1 else "")
    if not src:
        sys.exit("원천 워크북 경로가 필요합니다. SRC 환경변수 또는 첫 인자로 .xlsm 경로를 주세요.")
    src_path = Path(os.path.expanduser(src))
    if not src_path.exists():
        sys.exit(f"파일을 찾을 수 없습니다: {src_path}")

    print(f"읽는 중: {src_path}")
    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
    for need in (SHEET_PERSON, SHEET_CERT, SHEET_EDU):
        if need not in wb.sheetnames:
            sys.exit(f"필요한 시트가 없습니다: {need!r} (있는 시트: {wb.sheetnames})")

    emp_rows, name_to_id, emp_dups = load_employees(wb[SHEET_PERSON])
    cert_rows, _cert_map, cert_dups = load_certs(wb[SHEET_CERT])
    edu_rows, edu_missing = load_education(wb[SHEET_EDU], name_to_id)

    print("\n시드 CSV 생성 (Data/):")
    write_csv("01_직원.csv", H_EMPLOYEE, emp_rows)
    write_csv("02_자격증_마스터.csv", H_CERT, cert_rows)
    write_csv("04_직원_학력.csv", H_EDU, edu_rows)
    # --- 다음 단계용 빈 스텁(헤더만) — seed_from_csv() 가 6종을 모두 로드하므로 필요 ---
    write_csv("03_업무코드_마스터.csv", H_WORK, [])
    write_csv("05_직원_자격증.csv", H_EMP_CERT, [])
    write_csv("06_업무코드_자격증_매핑.csv", H_MAP, [])

    # --- 경고/리포트 ---
    print("\n요약:")
    print(f"  직원 {len(emp_rows)}명 · 자격증마스터 {len(cert_rows)}종 · 학력 {len(edu_rows)}건")
    if emp_dups:
        print(f"  [경고] 개인DB 동명이인(첫 항목만 매핑됨): {sorted(set(emp_dups))}")
    if cert_dups:
        print(f"  [경고] 자격증DB 중복 자격증명: {sorted(set(cert_dups))}")
    if edu_missing:
        miss = sorted(set(edu_missing))
        print(f"  [경고] 학력 {len(edu_missing)}건이 개인DB에 없는 이름 → 제외: {miss}")
    print("\n다음: repo 루트에서  python -c \"import db; db.seed_from_csv()\"  로 적재")


if __name__ == "__main__":
    main()
